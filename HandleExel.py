# -*- coding: utf-8 -*-
"""
Created on Fri Apr  7 16:09:56 2017

@author: roehl
"""
from openpyxl import Workbook
import numpy as np
import os
from QC_Functions import CheckFolder

def SaveMeasurements(AnalisisName,Cells):
    CheckFolder(AnalisisName)
    wb = Workbook()
    #wb.remove_sheet('Sheet')
    #create tabs
    tab_general_info = wb.create_sheet('General Info')
    
    tab_cell_based_measurements = wb.create_sheet('Cell Based Measurements')
    tab_cell_based_measurements['A1'] = 'Cell'
    tab_cell_based_measurements['B1'] = 'Area'
    tab_cell_based_measurements['C1'] = 'Perimeter'
    tab_cell_based_measurements['D1'] = 'NuclyOffset'
    tab_cell_based_measurements['F1'] = 'ConnectedCells'   
    
    tab_basic_measurements = wb.create_sheet('Basic Measurements')
    tab_basic_measurements['A1'] = ''
    tab_basic_measurements['B1'] = ''
    tab_basic_measurements['C1'] = 'Interface Area'
    tab_basic_measurements['D1'] = 'Interface Length'
    tab_basic_measurements['E1'] = 'Internuclear Distance'
    tab_basic_measurements['F1'] = 'Predicted Interface Length'
    tab_basic_measurements['G1'] = 'Summed Interface Area Intensity Label 1'
    tab_basic_measurements['H1'] = 'Summed Interface Area Intensity Label 2'
    tab_basic_measurements['I1'] = 'Fragmented Junction Length'
    tab_basic_measurements['J1'] = 'Junction Protein Area'
    tab_basic_measurements['K1'] = 'Predicted Fragmented Junction Length'
    tab_basic_measurements['L1'] = 'Summed Junction Area Intensity Label 1'
    tab_basic_measurements['M1'] = 'Summed Junction Area Intensity Label 2'
    tab_basic_measurements['N1'] = 'Junction Area'
    tab_basic_measurements['O1'] = 'Induvidual Junction Pieces'
    
    tab_derived_measurements = wb.create_sheet('Derived Measurements')
    tab_derived_measurements['A1'] = ''
    tab_derived_measurements['B1'] = ''    
    tab_derived_measurements['C1'] = 'Interface Linearity Index'
    tab_derived_measurements['D1'] = 'Coverage Index'  
    tab_derived_measurements['E1'] = 'Interface Occupancy'
    tab_derived_measurements['F1'] = 'Junction Label 1 Intensity Per Interface Area'  
    tab_derived_measurements['G1'] = 'Cluste rDensity'
    tab_derived_measurements['H1'] = ''  
    
    count = 2
    
    for i,k in enumerate(Cells.keys()):
        # Create General Cell Information
        tab_cell_based_measurements['A{}'.format(i+2)] = k
        tab_cell_based_measurements['B{}'.format(i+2)] = Cells[k]['Area']
        tab_cell_based_measurements['C{}'.format(i+2)] = Cells[k]['Perimeter']
        tab_cell_based_measurements['D{}'.format(i+2)] = Cells[k]['NuclyOffset']
        tab_cell_based_measurements['F{}'.format(i+2)] = str(Cells[k]['ConnectedCells'])
        
        #
        if len(Cells[k]['AnalisedJunctions']['processedConnections']) > 0:
            for com in Cells[k]['AnalisedJunctions']['processedConnections']:
                #Basic Measurements
                tab_basic_measurements['A{}'.format(count)] = k
                tab_basic_measurements['B{}'.format(count)] = com
                tab_basic_measurements['C{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['InterfaceArea']
                tab_basic_measurements['D{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['InterfaceLength']
                tab_basic_measurements['E{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['Internuclear Distance']
                tab_basic_measurements['F{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['PredictedInterfaceLength']
                tab_basic_measurements['G{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['SummedInterfaceAreaIntensityLabel1']
                tab_basic_measurements['H{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['SummedInterfaceAreaIntensityLabel2']
                tab_basic_measurements['I{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['FragmentedJunctionLength']
                tab_basic_measurements['J{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['JunctionProteinArea']
                tab_basic_measurements['K{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['PredictedFragmentedJunctionLength']
                tab_basic_measurements['L{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['SummedJunctionAreaIntensityLabel1']
                tab_basic_measurements['M{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['SummedJunctionAreaIntensityLabel2']
                tab_basic_measurements['N{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['JunctionArea']
                #tab_basic_measurements['O{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['']
                
                #Create Derived Measurements
                tab_derived_measurements['A{}'.format(count)] = k
                tab_derived_measurements['B{}'.format(count)] = com
                tab_derived_measurements['C{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['InterfaceLinearityIndex']
                tab_derived_measurements['D{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['CoverageIndex']
                tab_derived_measurements['E{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['InterfaceOccupancy']
                tab_derived_measurements['F{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['JunctionLabel1IntensityPerInterfaceArea']
                tab_derived_measurements['G{}'.format(count)] = Cells[k]['AnalisedJunctions'][com]['ClusterDensity']               
                
                count+=1
            

    
    wb.save(os.path.join('Results',AnalisisName,'Results','{}.xlsx'.format(AnalisisName)))   


